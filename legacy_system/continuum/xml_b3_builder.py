import openpyxl
from xml.dom import minidom
import pandas as pd
from os import listdir

xlfile = 'Application Tree.xlsx'

# <?xml version="1.0" encoding="utf-8"?>
# <ObjectSet ExportMode="Standard" Version="1.8.1.79" Note="TypesFirst">
#   <MetaInformation>
#     <ExportMode Value="Standard" />
#     <RuntimeVersion Value="1.8.1.79" />
#     <SourceVersion Value="1.8.1.79" />
#     <ServerFullPath Value="/Clives-ES/Servers/Clives-AS" />
#   </MetaInformation>
#   <ExportedObjects>
# 		<OI NAME="BACnet Analog Input (Continuum)" TYPE="bacnet.b3.point.analog.Input">
#			<PI Name="Channel" Value="1" />
#		</OI>
#		<OI NAME="BACnet Temp Input (Continuum)" TYPE="bacnet.b3.point.analog.Input">
#			<PI Name="Channel" Value="4" />
#			<PI Name="ElectricType" Value="4" />
#		</OI>
#
# "bacnet.b3.point.analog.Input"
#  - voltage (no tag)
#  - temperature <PI Name="ElectricType" Value="4" /> | "ElecType" 'ACCTemp(DEGC)'
# "bacnet.b3.point.analog.Output"
# "bacnet.b3.point.analog.Value"
# "bacnet.b3.point.digital.Input"
# "bacnet.b3.point.digital.Output"
# "bacnet.b3.point.digital.Value"
# "bacnet.b3.point.multistate.Input"
# "bacnet.b3.point.multistate.Value"
# "bacnet.b3.point.datetime.Value"
# "bacnet.b3.point.string.Value"

class b3ApplicationBuilder(object):

	def __init__(self, xmlfile=None, objects_xlfile=None):
		
		self.xmlfile = xmlfile
		self.objects_xlfile = objects_xlfile
		
		self.xml_head = """<?xml version="1.0" encoding="utf-8"?>
			<ObjectSet ExportMode="Standard" Version="1.8.1.79" Note="TypesFirst">
			\t<MetaInformation>
			\t\t<ExportMode Value="Standard" />
			\t\t<RuntimeVersion Value="1.8.1.79" />
			\t\t<SourceVersion Value="1.8.1.79" />
			\t\t<ServerFullPath Value="/Clives-ES/Servers/Clives-AS" />
			\t</MetaInformation>
			\t<ExportedObjects>"""

		self.xml_foot = """\t</ExportedObjects>
			</ObjectSet>"""

		self.object_types = {
			'alarms': [],
			'schedules': [],
			'variables': [
				"bacnet.b3.point.analog.Value", # analog
				"bacnet.b3.point.digital.Value", # digital
				"bacnet.b3.point.multistate.Value", # multi-state
				"bacnet.b3.point.datetime.Value", # timestamp
				"bacnet.b3.point.string.Value"  # string
			],
			"inputs": [
					"bacnet.b3.point.digital.Input",
					"bacnet.b3.point.multistate.Input",
					"bacnet.b3.point.analog.Input",
			],
			"outputs": [
					"bacnet.b3.point.analog.Output",
					"bacnet.b3.point.digital.Output",
					"bacnet.b3.point.multistate.Output"
			]
		}
		
		self.infinity_object_types = {
			"InfinityNumeric": "variables",
			"InfinityInput": "inputs",
			"InfinityOutput": "outputs",
			"InfinityString": "variables"
		}
		
	def close_element(self):
		return '</OI>'
	
	def create_object_element_by_name_type(self, object, infinity_object_type):
		"""
		Create XML element representing an SBO object.
		This element should be an end node with no children.
		eg <OI NAME="BACnet Analog Input (Continuum)" TYPE="bacnet.b3.point.analog.Input" />
		"""
		type = self.get_type_from_object(object, infinity_object_type)
		properties = self.get_object_properties(object, infinity_object_type)
		if properties is None:
			return '<OI NAME="' + object["name"] + '" TYPE="' + type + '" />'
		else:			
			element_open = '<OI NAME="' + object["name"] + '" TYPE="' + type + '" >'
			return element_open + properties + '\n' + self.close_element()
	
	def get_object_properties(self, object, infinity_object_type):
		properties = ""
		if infinity_object_type == "InfinityInput":
			properties += self.create_property_element_by_name_value("Channel", object["Channel"])
			if object["ElecType"] == str():
				if "ACCTemp" in object["ElecType"]:
					properties += self.create_property_element_by_name_value("ElectricType", "4")
		elif infinity_object_type == "InfinityOutput":
			properties += self.create_property_element_by_name_value("Channel", object["Channel"])
		else:
			properties = None
		return properties
	
	def get_type_from_object(self, object, infinity_object_type):
		object_type = self.infinity_object_types[infinity_object_type]
		
		if infinity_object_type == "InfinityString":
			type = "bacnet.b3.point.string.Value"
		elif infinity_object_type == "InfinityNumeric":
			if object["ElecType"] == "Digital":
				type = "bacnet.b3.point.digital.Value"
			else:
				type = "bacnet.b3.point.analog.Value"
		elif infinity_object_type == "InfinityInput":
			if object["ElecType"] == "Digital":
				type = "bacnet.b3.point.digital.Input"
			else:
				type = "bacnet.b3.point.analog.Input"
		elif infinity_object_type == "InfinityOutput":
			if object["ElecType"] == "Digital":
				type = "bacnet.b3.point.digital.Output"
			else:
				type = "bacnet.b3.point.analog.Output"
		else:
			type = ""
		return type
	
	def create_property_element_by_name_value(self, name, value):
		"""
		Create XML element representing an SBO object.
		This element should be an end node with no children.
		eg <PI Name="ElectricType" Value="4" />
		"""
		return '<PI Name="' + name + '" Value="' + str(value) + '" />'
	
	def create_folders_from_list(self, folder_names, children_dict={}):
		elements = str()
		for name in folder_names:
			if children_dict.get(name):
				children = children_dict[name]
			else:
				children = None
			elements += '\n' + self.create_folder_by_name(name, children)
		return elements

	def create_objects_from_excelsheet(self, sheetname, children=None, include_common=True):
		elements = str()
		df = pd.read_excel(self.objects_xlfile, sheetname=sheetname)
		for n, object in df.iterrows():
			print(object['name'])
			elements += '\n' + self.create_object_element_by_name_type(object, sheetname)
		return elements
	
	def get_object_subtype(self, object_type, subtype):
		for sbo_subtype in self.folder_object_types[object_type]:
			if sbo_subtype.split('.')[2] == subtype:
				return sbo_subtype
		
	def make_xml(self, write_result=True, print_result=False):
	
		xml_str = self.xml_head
		
		# xml_child_folders = self.create_folders_from_list(self.child_folder_names)
		wb = openpyxl.load_workbook(self.objects_xlfile)
		allsheetnames = wb.get_sheet_names()
		print(allsheetnames)
		sheetnames = [s for s in allsheetnames if s in self.infinity_object_types]
		print(sheetnames)
		for sheetname in sheetnames:
			print("\n" + sheetname)
			# make sure sheet not empty
			if wb[sheetname]['A1'].value is not None:
				xml_str += '\n' + self.create_objects_from_excelsheet(sheetname)
	
		xml_str += '\n' +self. xml_foot
		
		if print_result:
			print(xml_str)
		if write_result:
			with open(self.xmlfile, "w") as outfile:
				outfile.write(xml_str)
	

	
# define dump file locations
def get_xlfiles(path="."):
	files = listdir(path)
	xlfiles = [file for file in files if "ddc_objects" in file.lower() and ".xlsx" in file.lower()]
	return xlfiles

xlfiles = get_xlfiles()
	
for xlfile in xlfiles:
	print(xlfile)
	b3xmlbuilder = b3ApplicationBuilder(xmlfile=xlfile.split(".")[0]+".xml", objects_xlfile=xlfile)
	b3xmlbuilder.make_xml(write_result=True, print_result=True)


