import xlsxwriter
import openpyxl
from xml.dom import minidom
import pandas as pd
from os import listdir

# <?xml version="1.0" encoding="utf-8"?>
# <ObjectSet ExportMode="Standard" Version="1.8.1.79" Note="TypesFirst">
#   <MetaInformation>
#     <ExportMode Value="Standard" />
#     <RuntimeVersion Value="1.8.1.79" />
#     <SourceVersion Value="1.8.1.79" />
#     <ServerFullPath Value="/Clives-ES/Servers/Clives-AS" />
#   </MetaInformation>
#   <ExportedObjects>
# 		<OI NAME="BACnet Analog Input (Continuum)" TYPE="bacnet.b3.point.analog.Input">
#			<PI Name="Channel" Value="1" />
#		</OI>
#		<OI NAME="BACnet Temp Input (Continuum)" TYPE="bacnet.b3.point.analog.Input">
#			<PI Name="Channel" Value="4" />
#			<PI Name="ElectricType" Value="4" />
#		</OI>
#
# "bacnet.b3.point.analog.Input"
#  - voltage (no tag)
#  - temperature <PI Name="ElectricType" Value="4" /> | "ElecType" 'ACCTemp(DEGC)'
# "bacnet.b3.point.analog.Output"
# "bacnet.b3.point.analog.Value"
# "bacnet.b3.point.digital.Input"
# "bacnet.b3.point.digital.Output"
# "bacnet.b3.point.digital.Value"
# "bacnet.b3.point.multistate.Input"
# "bacnet.b3.point.multistate.Value"
# "bacnet.b3.point.datetime.Value"
# "bacnet.b3.point.string.Value"

class DmpfileExtractor(object):
	
	def __init__(self, dmpfile=None):
		self.b3_object_types = [
			"Report",
			"EventEnrollment",
			"EventNotification",
			"Schedule",
			"InfinityNumeric",
			"InfinityInput",
			"InfinityOutput",
			"InfinityString",
			"Group",
			"InfinityProgram",
			"InfinitySystemVariable"
		]
		self.dmpfile = dmpfile
		if self.dmpfile is not None:
			self.set_dmpfile(dmpfile)
		
	def set_dmpfile(self, dmpfile):
		self.dmpfile = dmpfile
		with open(dmpfile, 'rb') as f:
			# self.data = f.readlines()
			self.data = [line.strip() for line in f]
	
	def get_b3_objects(self, verbose=False):
		self.objects = {}
		for object_type in self.b3_object_types:
				self.objects[object_type] = self.get_b3_object_by_type(object_type)
		if verbose is True:
			self.get_b3_objects_attr()

	def get_b3_object_by_type(self, object_type):
		"""
		When passed the dump file as a list of rows,
		return a list of all objects that match a certain type
		"""
		objects = []
		for row in self.data:
			# check if row not empty
			if len(row.split()) > 0:
				# for object list, object type is first item after splitting whitespace
				if row.split()[0] == object_type:					
					objects.append(self.get_b3_object_name(row))
		return objects
		
	def get_b3_object_name(self, row):
		return row.split()[4]
		
	def get_b3_objects_attr(self):
		for object_type in self.objects:
			for object_name in self.objects[object_type]:
				self.get_b3_object_attr_by_name_type(object_name, object_type)
	
	def get_b3_object_attr_by_name_type(self, object_name, object_type):
		# print(object_name)
		object_attr = {"name": object_name}
		try:
			# get range of data representing the object attributes
			i1 = self.data.index("Object : " + object_name)
			i2 = i1 + self.data[i1:].index("EndObject")
		except ValueError:
			i1 = i2 = None
		if i1 is not None:
			if object_type in ["InfinityNumeric", "InfinityInput", "InfinityOutput"]:
				self.add_object_attr(object_attr, self.data[i1:i2], "ElecType")
				self.add_object_attr(object_attr, self.data[i1:i2], "Value")
				if object_type in ["InfinityInput", "InfinityOutput"]:
					self.add_object_attr(object_attr, self.data[i1:i2], "Channel")
				print(object_attr)
		self.objects[object_type][self.objects[object_type].index(object_name)] = object_attr
				
	def get_variable_subtype(self, object_type, is_digital):
		"""
		Returns a guess of the I/O type, eg BV (binary value), AO (analog output)
		based on the Infinity object type and whether or not the signal is digital.
		"""
		is_digital_switch = {
			True: "B",
			False: "A"
		}
		object_type_switch = {
			"InfinityNumeric": "V",
			"InfinityInput": "I",
			"InfinityOutput": "O"
		}
		return is_digital_switch[is_digital] + object_type_switch[object_type]
		
	
	def get_attr_value(self, object_data, attr_key):
		attr_value = ""
		for attr in object_data:
			if attr_key + " :" in attr:
				attr_value = attr.split(":")[-1].strip()
				break
		return attr_value
		
	def add_object_attr(self, object_attr, object_data, attr_key):
		object_attr[attr_key] = self.get_attr_value(object_data, attr_key)
		
	
	def to_excel(self, workbook="dmpfile_extraction.xlsx"):
		wb = xlsxwriter.Workbook(workbook)
		for object_type in self.b3_object_types:
			sheet = wb.add_worksheet(object_type)
			self.write_objects_to_sheet(object_type, sheet)
		wb.close()
	
	def write_objects_to_sheet(self, object_type, sheet):
		if len(self.objects[object_type]) < 1:
			pass
		else:
			headers = self.objects[object_type][0].keys()
			first_row = 0
			# write headers
			for header in headers:
				col = headers.index(header) # we are keeping order.
				sheet.write(0, col, header) # we have written first row which is the header of worksheet also.

			for n, object in enumerate(self.objects[object_type]):
				for header, value in object.items():
					col = headers.index(header)
					sheet.write(n+1, col, value)

					
class b3ApplicationBuilder(object):

	def __init__(self, xmlfile=None, objects_xlfile=None):
		
		self.xmlfile = xmlfile
		self.objects_xlfile = objects_xlfile
		
		self.xml_head = """<?xml version="1.0" encoding="utf-8"?>
			<ObjectSet ExportMode="Standard" Version="1.8.1.79" Note="TypesFirst">
			\t<MetaInformation>
			\t\t<ExportMode Value="Standard" />
			\t\t<RuntimeVersion Value="1.8.1.79" />
			\t\t<SourceVersion Value="1.8.1.79" />
			\t\t<ServerFullPath Value="/Clives-ES/Servers/Clives-AS" />
			\t</MetaInformation>
			\t<ExportedObjects>"""

		self.xml_foot = """\t</ExportedObjects>
			</ObjectSet>"""

		self.object_types = {
			'alarms': [],
			'schedules': [],
			'variables': [
				"bacnet.b3.point.analog.Value", # analog
				"bacnet.b3.point.digital.Value", # digital
				"bacnet.b3.point.multistate.Value", # multi-state
				"bacnet.b3.point.datetime.Value", # timestamp
				"bacnet.b3.point.string.Value"  # string
			],
			"inputs": [
					"bacnet.b3.point.digital.Input",
					"bacnet.b3.point.multistate.Input",
					"bacnet.b3.point.analog.Input",
			],
			"outputs": [
					"bacnet.b3.point.analog.Output",
					"bacnet.b3.point.digital.Output",
					"bacnet.b3.point.multistate.Output"
			]
		}
		
		self.infinity_object_types = {
			"InfinityNumeric": "variables",
			"InfinityInput": "inputs",
			"InfinityOutput": "outputs",
			"InfinityString": "variables"
		}
		
	def close_element(self):
		return '</OI>'
	
	def create_object_element_by_name_type(self, object, infinity_object_type):
		"""
		Create XML element representing an SBO object.
		This element should be an end node with no children.
		eg <OI NAME="BACnet Analog Input (Continuum)" TYPE="bacnet.b3.point.analog.Input" />
		"""
		type = self.get_type_from_object(object, infinity_object_type)
		properties = self.get_object_properties(object, infinity_object_type)
		if properties is None:
			return '<OI NAME="' + object["name"] + '" TYPE="' + type + '" />'
		else:			
			element_open = '<OI NAME="' + object["name"] + '" TYPE="' + type + '" >'
			return element_open + properties + '\n' + self.close_element()
	
	def get_object_properties(self, object, infinity_object_type):
		properties = ""
		if infinity_object_type == "InfinityInput":
			properties += self.create_property_element_by_name_value("Channel", object["Channel"])
			if object["ElecType"] == str():
				if "ACCTemp" in object["ElecType"]:
					properties += self.create_property_element_by_name_value("ElectricType", "4")
		elif infinity_object_type == "InfinityOutput":
			properties += self.create_property_element_by_name_value("Channel", object["Channel"])
		else:
			properties = None
		return properties
	
	def get_type_from_object(self, object, infinity_object_type):
		object_type = self.infinity_object_types[infinity_object_type]
		
		if infinity_object_type == "InfinityString":
			type = "bacnet.b3.point.string.Value"
		elif infinity_object_type == "InfinityNumeric":
			if object["ElecType"] == "Digital":
				type = "bacnet.b3.point.digital.Value"
			else:
				type = "bacnet.b3.point.analog.Value"
		elif infinity_object_type == "InfinityInput":
			if object["ElecType"] == "Digital":
				type = "bacnet.b3.point.digital.Input"
			else:
				type = "bacnet.b3.point.analog.Input"
		elif infinity_object_type == "InfinityOutput":
			if object["ElecType"] == "Digital":
				type = "bacnet.b3.point.digital.Output"
			else:
				type = "bacnet.b3.point.analog.Output"
		else:
			type = ""
		return type
	
	def create_property_element_by_name_value(self, name, value):
		"""
		Create XML element representing an SBO object.
		This element should be an end node with no children.
		eg <PI Name="ElectricType" Value="4" />
		"""
		return '<PI Name="' + name + '" Value="' + str(value) + '" />'
	
	def create_folders_from_list(self, folder_names, children_dict={}):
		elements = str()
		for name in folder_names:
			if children_dict.get(name):
				children = children_dict[name]
			else:
				children = None
			elements += '\n' + self.create_folder_by_name(name, children)
		return elements

	def create_objects_from_excelsheet(self, sheetname, children=None, include_common=True):
		elements = str()
		df = pd.read_excel(self.objects_xlfile, sheetname=sheetname)
		for n, object in df.iterrows():
			print(object['name'])
			elements += '\n' + self.create_object_element_by_name_type(object, sheetname)
		return elements
	
	def get_object_subtype(self, object_type, subtype):
		for sbo_subtype in self.folder_object_types[object_type]:
			if sbo_subtype.split('.')[2] == subtype:
				return sbo_subtype
		
	def make_xml(self, write_result=True, print_result=False):
	
		xml_str = self.xml_head
		
		# xml_child_folders = self.create_folders_from_list(self.child_folder_names)
		wb = openpyxl.load_workbook(self.objects_xlfile)
		allsheetnames = wb.get_sheet_names()
		print(allsheetnames)
		sheetnames = [s for s in allsheetnames if s in self.infinity_object_types]
		print(sheetnames)
		for sheetname in sheetnames:
			print("\n" + sheetname)
			# make sure sheet not empty
			if wb[sheetname]['A1'].value is not None:
				xml_str += '\n' + self.create_objects_from_excelsheet(sheetname)
	
		xml_str += '\n' +self. xml_foot
		
		if print_result:
			print(xml_str)
		if write_result:
			with open(self.xmlfile, "w") as outfile:
				outfile.write(xml_str)

if __name__ == '__main__':
	# Example DmpfileExtractor usage
	def get_dmpfiles(path="."):
		files = listdir(path)
		dmpfiles = [file for file in files if ".dmp" in file.lower()]
		return dmpfiles
		
	dumpfiles = get_dmpfiles()
	print(dumpfiles)

	for dumpfile in dumpfiles:
		xl_outfile = 'ddc_objects_' + dumpfile.split(".")[0] + '_.xlsx'
		extractor = DmpfileExtractor(dmpfile=dumpfile)
		extractor.get_b3_objects(verbose=True)
		extractor.to_excel(workbook=xl_outfile)
	
	# Example b3ApplicationBuilder usage
	def get_xlfiles(path="."):
		files = listdir(path)
		xlfiles = [file for file in files if "ddc_objects" in file.lower() and ".xlsx" in file.lower()]
		return xlfiles

	xlfiles = get_xlfiles()
	
	for xlfile in xlfiles:
		print(xlfile)
		b3xmlbuilder = b3ApplicationBuilder(xmlfile=xlfile.split(".")[0]+".xml", objects_xlfile=xlfile)
		b3xmlbuilder.make_xml(write_result=True, print_result=True)

