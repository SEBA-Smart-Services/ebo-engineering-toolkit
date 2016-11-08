import xlsxwriter
import pandas as pd
from openpyxl import load_workbook
import json
from io import BytesIO


def module_io(module, module_type):
	"""
	When passed a module (row from df)
	and the module type (dict from config json) 
	return a list of IO addresses
	"""
	if module_type['name'] == module['Type']:
		io_addresses = get_io_addresses(module_type)
	return io_addresses

def get_io_addresses(module_type):
	"""
	When passed module type (dict from config json) 
	return a list of IO addresses
	"""
	io_addresses = []
	for io_type in ['DI', 'UI', 'DO', 'AO']:
		for i in range(int(module_type[io_type + '_count'])):
			io_addresses.append(io_type + str(i+1).zfill(2))
	return(io_addresses)

def replace_key_value(d, key, new_value):
	d_new = d
	for k, v in d_new.items():
		if k == key:
			d_new[k] = new_value
	return d_new
	
def create_module_df(module_cols, io_addresses):
	"""
	When passes the list of IO module columns
	and list of IO addresses
	return a new IO module dataframe
	"""
	empty_col = ['']*len(io_addresses)
	module_data = dict((col, empty_col) for col in module_cols)
	module_data = replace_key_value(module_data, "IO address", io_addresses)
	module_df = pd.DataFrame(module_data)
	module_df = module_df[module_cols]
	return module_df

def format_xlsheet(writer, module_name):
	worksheet = writer.sheets[module_name]
	format = workbook.add_format({'font_name': 'Consolas', 'font_size': 10})
	
	worksheet.set_column('A:A', 10)
	worksheet.set_column('B:B', 24)
	worksheet.set_column('C:C', 4)
	worksheet.set_column('D:F', 32)
	worksheet.set_column('G:I', 8)
	# print format
	worksheet.set_landscape()
	worksheet.set_paper(9)
	worksheet.center_horizontally()
	worksheet.fit_to_pages(1, 1)
	worksheet.set_margins(top=1)
	image_file = open(logofile, 'rb')
	image_data = BytesIO(image_file.read())
	worksheet.set_header('&L&[Picture]&CController: bla\nModule: &A&RProject number: ' + project_number + '\nProject name: ' + project_name,
		{'image_left': logofile,
		'image_data_left': image_data})
	
	worksheet.set_footer('&CPage &P of &N')

# job info
project_number = '2353597Q'
project_name = 'Cbus SBO Chiller Cutover'

# filenames
xl_infile = 'test_IO_bus.xlsx'
xl_outfile = 'test_IO_bus.xlsx'
moduletypes_jsonfile = 'module_types.json'
logofile = 'logo.png'

# xl worksheet with list of modules and addresses
modules_ws = 'Module Summary'

# list of IO module columns
module_cols = [
	"IO address",
	"Name",
	"Type",
	"Description",
	"Field device",
	"Location",
	"Continuum Controller",
	"Continuum IO address",
	"Wire number",
]

# read module types json as dict
with open(moduletypes_jsonfile, 'rb') as f:
	moduletypes = json.loads(f.read())

# wb = load_workbook(filename = xl_infile)

# Create a Pandas Excel writer using XlsxWriter as the engine.
writer = pd.ExcelWriter(xl_outfile, engine='xlsxwriter')
workbook = writer.book

# create a modules df from xl worksheet
modules_df = pd.read_excel(xl_infile, sheetname=modules_ws)
# write module summary to new xl workbook
modules_df.to_excel(writer, sheet_name=modules_ws, index=False)


# IO module columns to dict for df creation
module_schema = dict((col,[]) for col in module_cols)

for n, module in modules_df.iterrows():
	# skip address 1 & 2 to get to IO modules
	if n < 2:
		continue
	print(str(module['Address']) + " " + str(module['Name']))
	# get list of IO addresses for this module
	for n, type in enumerate(moduletypes):
		if type['name'] == module['Type']:
			print(module)
			io_addresses = module_io(module, type)
	# convert IO address list to dataframe data
	# create dataframe data for empty IO module with IO addresses only
	module_df = create_module_df(module_cols, io_addresses)
	print(module_df)
	module_df.to_excel(writer, sheet_name=module['Name'], index=False)
	format_xlsheet(writer, module['Name'])

# Close the Pandas Excel writer and output the Excel file.
writer.save()


# workbook = xlsxwriter.Workbook(xlfile)
# worksheet1 = workbook.add_worksheet('Penis')

# worksheet.write('A1', 'Hello')

# workbook.close()