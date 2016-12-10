################################
# modbus_slave_transition.py
# Clive Gross
# 10 Dec 2016
################################
#
################
# LICENSE
################
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.

# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.

# You should have received a copy of the GNU General Public License
# along with this program.  If not, see <http://www.gnu.org/licenses/>.
#
################
# DESCRIPTION
################
# Forms part of the Vista to SBO transition toolkit
#
# Converts a Vista Modbus slave registers to SBO Modbus Interface registers.
# ie:
# 	- AS:
#		- Modbus Interface:
#			- Local Register:
#				- Name: CH1 Fault
#				- Register number: 40002
#			- Local Register:
#				- Name: CH1 Status
#				- Register number: 40003
#
# Requires manually copying the XBuilder Device Editor table to an Excel spreadsheet.
#
# NOTE: will not preserve any bindings of the modbus registers. These will have to manually rebound
# NOTE: currently only works for Digital Coil register types. Other types can be easily added.
#
################################
# Example Excel file for reading
################################
# 		A:			B:			C:				D:				E...
# (		object_name	not_used	register_number	register_type 	not_used)
# 1:	CH1 Fault				40002			16 bit Unsigned	
# 2:	CH1 Status				40003			16 bit Unsigned	
#
#
################################
# Example XML for importing into SBO
################################
# <?xml version="1.0" encoding="utf-8"?>
# <ObjectSet ExportMode="Standard" Version="1.8.1.79" Note="TypesFirst">
  # <MetaInformation>
    # <ExportMode Value="Standard" />
    # <RuntimeVersion Value="1.8.1.79" />
    # <SourceVersion Value="1.8.1.79" />
    # <ServerFullPath Value="/KGSB-ES/Servers/Services-AS001" />
  # </MetaInformation>
  # <ExportedObjects>
    # <OI NAME="Modbus Interface" TYPE="modbus.network.SlaveDevice">
      # <OI NAME="CH1 Fault" TYPE="modbus.point.BinaryValue">
        # <PI Name="RegisterNumber" Value="40002" />
      # </OI>
      # <OI NAME="CH1 Status" TYPE="modbus.point.BinaryValue">
        # <PI Name="RegisterNumber" Value="40003" />
      # </OI>	  
    # </OI>
  # </ExportedObjects>
# </ObjectSet>
#
#
import openpyxl
from xml.dom import minidom


class ModbusSlaveTransition(object):

	def __init__(self, xlfile=None,  xmlfile=None):
		self.xlfile = xlfile
		self.xmlfile = xmlfile
		# initialise list of Modbus objects
		self.modbus_objects = []
		self.modbus_interface_name = "Modbus Interface"
		
		self.xml_head = """<?xml version="1.0" encoding="utf-8"?>
			<ObjectSet ExportMode="Standard" Version="1.8.1.79" Note="TypesFirst">
			\t<MetaInformation>
			\t\t<ExportMode Value="Standard" />
			\t\t<RuntimeVersion Value="1.8.1.79" />
			\t\t<SourceVersion Value="1.8.1.79" />
			\t\t<ServerFullPath Value="/SmartStruxure Server" />
			\t</MetaInformation>
			\t<ExportedObjects>"""

		self.xml_foot = """\t</ExportedObjects>
			</ObjectSet>"""
		
	def set_xlfile(self, xlfile):
		"""
		Set the Excel file name to read containing the Modbus registers.
		"""	
		self.xlfile = xlfile
	
	def set_xmlfile(self, xmlfile):
		"""
		Set the file name to output the XML for importing into SBO.
		"""
		self.xmlfile = xmlfile
		
	def clear_modbus_objects(self):
		"""
		empties self.modbus_objects
		"""
		self.modbus_objects = []
	
	def objects_to_dict(self, name, register_number):
		return \
			{
				'name': name,
				'register_number': register_number
			}
			
	def close_element(self, element_type):
		return '</' + self.lookup_tag(element_type) + '>'
	
	def create_element_by_name_type(self, name, type, children=None, element_type='object'):
		"""
		Create XML element representing an SBO object.
		If the element is an end node with no children, the tag will be closed.
		If the element has children, nest the children, then close the tag.
		If the element xml type is 'parameter', use <PI ... /> instead of <OI ... />
		"""
		element_base = self.make_element(element_type)
		element_base = element_base.replace('{{ name }}', name)
		element_base = element_base.replace('{{ type }}', type)
		if children:
			element_open = element_base.replace(' />', ' >')
			return element_open + '\n' + children + '\n' + self.close_element(element_type)
		else:
			return element_base
			
	def make_element_by_attrs(self, attributes, children=None):
		"""
		Create XML element representing an SBO object or parameter.
		If the element is an end node with no children, the tag will be closed.
		If the element has children, nest the children, then close the tag.
		If the element xml type is 'parameter', use <PI ... /> instead of <OI ... />
		"""
		element = self.make_element(attributes['tag'])
		for attr_name, attr_value in attributes.iteritems():
			if attr_name != 'tag':
				element = element.replace('{{ ' + attr_name + ' }}', attr_value)
		if children:
			element_open = element.replace(' />', ' >')
			element = element_open + '\n' + children + '\n' + self.close_element(attributes['tag'])
		return element
				
	def lookup_tag(self, element_type):
		"""
		Returns the tag name for opening and closing an element based on the element_type
		"""
		tag_lookup_dict = {
			'object': 'OI',
			'parameter': 'PI'
		}
		return tag_lookup_dict[element_type]
	
	def make_element(self, element_type):
		"""
		THIS NEEDS WORK
		this method prepares an xml element based on the element type
		including attributes. The attribute values are left with placeholders
		"""
		element_attribute_base = '{{ ATTR_NAME }}="{{ ATTR_VALUE }}"'
		element_base = '<{{ TAG }} />'
		element_base = element_base.replace('{{ TAG }}', self.lookup_tag(element_type))
		if element_type == 'object':
			attributes = {
								'NAME': '{{ name }}',
								'TYPE': '{{ type }}'
			}
		elif element_type == 'parameter':
			attributes = {
								'Name': '{{ name }}',
								'Value': '{{ value }}'
			}
		else:
			attributes = {}
		for attr_name, attr_value in attributes.iteritems():
			element_base = element_base.replace('/>', element_attribute_base + ' />')
			element_base = element_base.replace('{{ ATTR_NAME }}', attr_name)
			element_base = element_base.replace('{{ ATTR_VALUE }}', attr_value)
		return element_base		
			
	def xml_from_modbus_objects(self):
		"""
		NEEDS DECOUPLING
		loops through the list of modbus objects (self.modbus_objects, dict )
		and creates the xml element and any nested parameter elements
		appends them to the modbus objects xml
		"""
		xml = ''
		for register in self.modbus_objects:
			# create RegisterNumber parameter element
			if register['register_number'] != 'None':
				register_number_attr = {
					'tag': 'parameter',
					'name': 'RegisterNumber',
					'value': register['register_number']
				}
				xml_register_number = self.make_element_by_attrs(register_number_attr)
			else:
				xml_register_number = None
			# create Modbus register object element, nesting RegisterNumber parameter element
			register_attr = {
					'tag': 'object',
					'name': register['name'],
					'type': 'modbus.point.BinaryValue'
				}
			xml_register = self.make_element_by_attrs(register_attr, children=xml_register_number)
			xml += '\n' + xml_register
		return xml
	
	def xl_to_list(self):
		"""
		Read in the Excel worksbook containing registers.
		Read each sheet and try and convert items into list of dicts in the following format:
			[
				{
					'name': 'CH1 Fault',
					'register': '40002'
				},
				{
					'name': 'CH1 Status',
					'register': '40003'
				},
		"""
		# initialise workbook
		workbook = openpyxl.load_workbook(self.xlfile)
		# get sheets
		sheetnames = workbook.get_sheet_names()
		
		# iterate through sheets and rows inserting modbus registers into list
		for sheetname in sheetnames:
			sheet = workbook.get_sheet_by_name(sheetname)
			for row in sheet.iter_rows():
				register_dict = self.objects_to_dict(str(row[0].value), str(row[2].value))
				self.modbus_objects.append(register_dict)
	
	def make_xml(self, write_result=True, print_result=False):
		"""
		Generate XML and write to filename defined by self.xmlfile.
		This is the XML file that is imported into SBO
		"""
		# initialise xml content with xml header
		xml_str = self.xml_head
		
		# get list of Modbus register objects
		self.xl_to_list()
		
		# create xml for modbus objects
		xml_modbus_registers = self.xml_from_modbus_objects()
		
		# create xml for modbus interface
		xml_modbus_interface = self.create_element_by_name_type(self.modbus_interface_name, "modbus.network.SlaveDevice", children=xml_modbus_registers)
		
		xml_str += '\n' + xml_modbus_interface
		
		# complete xml content with footer
		xml_str += '\n' +self.xml_foot
		
		if print_result:
			print(xml_str)
		if write_result:
			with open(xmlfile, "w") as outfile:
				outfile.write(xml_str)


# EXECUTE
if __name__ == "__main__":
	# declare filenames/paths here
	xlfile = 'Example Modbus register schedule.xlsx'
	xmlfile = 'Example SBO Modbus registers output.xml'

	# instantiate ModbusSlaveTransition object
	sbo_modbus_converter = ModbusSlaveTransition(xlfile=xlfile, xmlfile=xmlfile)
	
	# create SBO xml for importing
	sbo_modbus_converter.make_xml()
	


