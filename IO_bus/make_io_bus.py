# IO module (NAME, TYPE)
# <OI NAME="AO-8" TYPE="io.AO8"></OI>
# 	properties:
#		ModuleID (Name, Value)
#		<PI Name="ModuleID" Value="5" />
#	objects:
#		IO (see IO)
#
# IO (DESCR, NAME, TYPE)
# <OI DESCR="chilled water valve" NAME="chilled water" TYPE="io.point.VoltageOutput">
# <OI DESCR="humidity" NAME="Resistive Input" TYPE="io.point.ResistiveInput">
#	properties:
#		LabelText (Name, Value)
#		<PI Name="LabelText" Value="0301" />
#		Value (Name, RefNull, Unit)
#		<PI Name="Value" RefNull="1" Unit="0x200001" />
#
#		[analog inputs and analog outputs]
#		ElecTopOfScale (Name, Value)
#		<PI Name="ElecTopOfScale" Value="10" />
#		EngBottomOfScale, EngTopOfScale (Name, Unit, Value)
#		<PI Name="EngBottomOfScale" Unit="0x200001" Value="0" />
#		<PI Name="EngTopOfScale" Unit="0x200001" Value="100" />
#
#		[outputs]
#		OutputChannelNumber (Name, Value)
#		<PI Name="OutputChannelNumber" Value="1" />
#
#		[inputs]
#		InputChannelNumber (Name, Value)
#		<PI Name="InputChannelNumber" Value="4" />
#
#
# {
#   "type": "io",
#	"
# }	

import openpyxl
import pandas as pd
from copy import deepcopy
import json
from lxml import etree

class SBO_xml_maker(object):

	def __init__(self, xmlfile=None, version="1.8.1.79"):
		
		self.xmlfile = xmlfile
		
		xml_head_template = """<?xml version="1.0" encoding="utf-8"?>
			<ObjectSet ExportMode="Standard" Version="{{ version }}" Note="TypesFirst">
			<MetaInformation>
			<ExportMode Value="Standard" />
			<RuntimeVersion Value="{{ version }}" />
			<SourceVersion Value="{{ version }}" />
			<ServerFullPath Value="/Clives-ES/Servers/Clives-AS" />
			</MetaInformation>
			<ExportedObjects>"""
		
		self.xml_head = xml_head_template.replace("{{ version }}", str(version))
		
		self.xml_foot = """</ExportedObjects>
			</ObjectSet>"""
			
	def set_xmlfile(self, xmlfile):
		self.xmlfile = xmlfile
	
	def close_element(self):
		return '</OI>'
	
	def create_object_element_by_name_type(self, object):
		"""
		Create XML element representing an SBO object.
		This element could be an end node with no children or
		could be an element containing nested properties.
		eg <OI NAME="BACnet Analog Input (Continuum)" TYPE="bacnet.b3.point.analog.Input" />
		eg 	<OI DESCR="temperature" NAME="Temperature Input" TYPE="io.point.TemperatureInput">
				<PI Name="InputChannelNumber" Value="1" />
				<PI Name="LabelText" Value="0401" />
			</OI>
		expects argument 'object' to be a dict conatining object data and properties data, see comments.
		"""
		properties = self.create_properties_from_object(object)
		# check if object has nested array of objects then run this method recursively
		if "objects" in object:
			nested_object_elements = ''
			for nested_object in object["objects"]:
				nested_object_elements += self.create_object_element_by_name_type(nested_object)
		else:
			nested_object_elements = None
		# create object element
		element = '<OI NAME="' + object["name"] + '" TYPE="' + object["type"] + '" DESCR="' + object["description"] + '" />'
		# nest properties if any inside element
		if properties is not None:
			element_open = element.replace('/>', '>')
			element = element_open + properties + self.close_element()
		# nest objects if any inside element
		if nested_object_elements is not None:
			element = element.replace(self.close_element(), nested_object_elements + self.close_element())
		return element
			
	def create_properties_from_object(self, object):
		"""
		Create a series of XML elements representing an SBO objects properties.
		These elements should be an end node with no children.
		eg	<PI Name="InputChannelNumber" Value="1" />
			<PI Name="LabelText" Value="0401" />
		"""
		properties_xml = str()
		for property in object["properties"]:
			element = self.create_property_element_by_name_value(property)
			properties_xml += element
		return properties_xml		

	def create_property_element_by_name_value(self, property):
		"""
		Create XML element representing an SBO object.
		This element should be an end node with no children.
		eg <PI Name="ElectricType" Value="4" />
		"""
		if len(property) == 0:
			return None
		elif len(property) == 1:
			name = property.keys()[0]
			value = property.values()[0]
			return '<PI Name="' + str(name) + '" Value="' + str(value) + '" />'
			

class SBO_io_bus_maker(SBO_xml_maker):
	
	def set_objects(self, objects):
		self.objects = objects
	
	## THIS IS GENERIC, MOVE TO PARENT CLASS
	def make_xml(self, print_only=False):
		xml_str = self.xml_head
		for module in self.objects:
			xml_str += self.create_object_element_by_name_type(module)
		# pretty print xml
		xml_str += self.xml_foot
		xml_str = xml_str.replace('\n', '').replace('\t', '')
		root = etree.fromstring(xml_str)
		xml_str = etree.tostring(root, pretty_print=True)
		if print_only:
			print(xml_str)
		else:
			with open(self.xmlfile, "w") as outfile:
				outfile.write(xml_str)
		

class SBO_points_list_reader(object):
	"""
	Give me an excel file and i will give you a json/dict.
	Turns an Excel points list into a points dictionary.
	Each sheet in the workbook is an IO module, each row is a point.
	The top row of each worksheet is the header.
	Check the class methods to see which columkn headers it is expecting.
	"""
	
	def __init__(self, datafile=None, ignore_sheets=['meta', 'settings', 'summary']):
		self.datafile = datafile
		self.ignore_sheets = ignore_sheets
		self.objects = []
		self.module_template = {"type": "io", "description": "", "properties": []}
		self.point_template = {"type": "io.point", "description": "", "properties": []}
		
	def set_datafile(self, datafile):
		self.datafile = datafile
		
	def read_io_modules(self):
		"""
		sheet name must be in format
		Mxx yyyyy or Mxx:yyyyyy or Mxx-yyy or Mxx_yyyy
		where xx is the module ID and yyy/yyyy/yyyyy/yyyyyy is the module type
		eg M05 AO-8
		eg M03:DO-FA-12
		"""
		wb = openpyxl.load_workbook(self.datafile)
		
		sheetnames = [s for s in wb.get_sheet_names() if s.lower() not in self.ignore_sheets]
		# loop through sheetnames as IO module names and store data
		for sheetname in sheetnames:
			module = deepcopy(self.module_template)
			address = int(sheetname[1:3])
			module["name"] = str(sheetname)
			module["type"] += '.' + self.reduce_str(sheetname[3:])
			module["properties"].append(
				{
					"ModuleID": str(address)
				}
			)
			module["objects"] = self.read_points(sheetname)
			self.objects.append(module)
			
	def read_points(self, sheetname):
		points_df = pd.read_excel(self.datafile, sheetname=sheetname)
		points_list = []
		for n, point in points_df.iterrows():
			# check this is not an empty address
			if 'empty' not in str(point['type']).lower():
				points_list.append(self.read_point(point))
		return points_list
		
	def read_point(self, point):
		point_dict = deepcopy(self.point_template)
		# set type
		point_dict['type'] += '.' + self.reduce_str(point["type"])
		# set description
		point_dict['description'] = str(point["description"])
		# set name
		point_dict['name'] = str(point["name"])
		# set input/output channel number property
		if "output" in point["type"].lower():
			point_dict["properties"].append(
				{
					"OutputChannelNumber": str(point["ch#"])
				}
			)
		elif "input" in point["type"].lower():
			point_dict["properties"].append(
				{
					"InputChannelNumber": str(point["ch#"])
				}
			)
		# set label property
		point_dict["properties"].append(
			{
				"LabelText": str(point["wire #"])
			}
		)

		return point_dict
		
	def reduce_str(self, phrase, lower=False):
		"""
		strip out whitepace and other nuisance characters that are dopped in SBO xml.
		optionally reduce to lower case
		"""
		reduced = phrase.replace(" ", "").replace("-", "").replace("_", "").replace(":", "")
		if lower:
			return str(reduced.lower())
		else:
			return str(reduced)
			
	def to_json(self, jsonfile='objects.json', print_only=False):
		"""
		dumps objects dictionary to json file or prints to screen
		"""
		json_str = json.dumps(self.objects, indent=4, sort_keys=True)
		if print_only:
			print(json_str)
		else:
			with open(jsonfile, 'w') as f:
				f.write(json_str)


if __name__ == '__main__':
	# set xl points list file name and SBO xml 	 file name
	xlfile = '2785970Q Points List R7.xlsx'
	xmlfile = xlfile.replace('.xlsx', '.xml')

	# read xl file and create IO bus objects
	# creates a dict for IO bus maker
	points_list_reader = SBO_points_list_reader(datafile=xlfile)
	points_list_reader.read_io_modules()
	points_list_reader.to_json(print_only=True)
	io_bus = points_list_reader.objects

	# convert objects to xml
	io_bus_maker = SBO_io_bus_maker(xmlfile=xmlfile)

	io_bus_maker.set_objects(io_bus)
	io_bus_maker.make_xml()
