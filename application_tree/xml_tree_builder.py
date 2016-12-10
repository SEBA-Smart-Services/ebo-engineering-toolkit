import openpyxl
from xml.dom import minidom

# <?xml version="1.0" encoding="utf-8"?>
# <ObjectSet ExportMode="Standard" Version="1.8.1.79" Note="TypesFirst">
#   <MetaInformation>
#     <ExportMode Value="Standard" />
#     <RuntimeVersion Value="1.8.1.79" />
#     <SourceVersion Value="1.8.1.79" />
#     <ServerFullPath Value="/SmartStruxure Server" />
#   </MetaInformation>
#   <ExportedObjects>
#     <OI NAME="one" TYPE="system.base.Folder">
#       <OI NAME="two" TYPE="system.base.Folder" />
#       <OI NAME="three" TYPE="system.base.Folder" />
#			<OI NAME="TestSchedule" TYPE="schedule.NSPDigitalSchedule" />
#			<OI NAME="Analog Value" TYPE="server.point.AV" />
#			<OI NAME="Digital Value" TYPE="server.point.BV" />
#			<OI NAME="Change of State Alarm" TYPE="alarm.ChangeOfStateAlarm" />
#		</OI>
#     </OI>
#   </ExportedObjects>
# </ObjectSet>

class ApplicationTreeBuilder(object):

	def __init__(self, xlfile=None,  xmlfile=None, commonobjects_xlfile=None):
		self.xlfile = xlfile
		self.xmlfile = xmlfile
		self.commonobjects_xlfile = commonobjects_xlfile
		
		self.xml_head = """<?xml version="1.0" encoding="utf-8"?>
			<ObjectSet ExportMode="Standard" Version="1.8.1.79" Note="TypesFirst">
			\t<MetaInformation>
			\t\t<ExportMode Value="Standard" />
			\t\t<RuntimeVersion Value="1.8.1.79" />
			\t\t<SourceVersion Value="1.8.1.79" />
			\t\t<ServerFullPath Value="/SmartStruxure Server" />
			\t</MetaInformation>
			\t<ExportedObjects>"""

		self.xml_foot = """\t</ExportedObjects>
			</ObjectSet>"""

		self.child_folder_names = [
			'Alarms',
			'Commissioning',
			'Documents',
			'Graphics',
			'Programs',
			'Schedules',
			'Trends',
			'Variables'
		]
	
		self.folder_object_types = {
			'Alarms': "alarm.ChangeOfStateAlarm",
			'Schedules': "schedule.NSPDigitalSchedule",
			'Variables': [
				"server.point.AV", # analog
				"server.point.BV", # digital
				"server.point.IV", # multi-state
				"server.point.SV", # string
				"server.point.TS"  # timestamp
			]			
		}
		
	def close_element(self):
		return '</OI>'
	
	def create_element_by_name_type(self, name, type):
		"""
		Create XML element representing an SBO object.
		This element should be an end node with no children.
		"""
		return '<OI NAME="' + name + '" TYPE="' + type + '" />'
	
	def create_common_grandchildren(self):
		pass
	
	def create_folder_by_name(self, name, children=None):
		"""
		Create XML element representing a folder.
		If folder contains child elements, nest child elements 'children' XML string
		If folder contains grandchild elements, nest grandchild elements XML in matching child XML		
		"""
		if children is not None:
			element_open = '<OI NAME="' + name + '" TYPE="system.base.Folder">'
			return element_open + children + '\n' + self.close_element()
		else:
			return '<OI NAME="' + name + '" TYPE="system.base.Folder" />'

	def create_folders_from_list(self, folder_names, children_dict={}):
		elements = str()
		for name in folder_names:
			if children_dict.get(name):
				children = children_dict[name]
			else:
				children = None
			elements += '\n' + self.create_folder_by_name(name, children)
		return elements

	def create_folders_from_excelsheet(self, sheet, children=None, include_common=True):
		elements = str()
		if include_common:
			if self.commonobjects_xlfile is not None:
				grandchildren = self.create_common_grandchildren()
			else:
				grandchildren = None
			elements += '\n' + self.create_folder_by_name('_Common', children)
		for cell in sheet.columns[0]:
			if cell.value is not None:
				elements += '\n' + self.create_folder_by_name(cell.value, children)
		return elements
	
	def get_object_subtype(self, object_type, subtype):
		for sbo_subtype in self.folder_object_types[object_type]:
			if sbo_subtype.split('.')[2] == subtype:
				return sbo_subtype
		
	def create_objects_from_excelbook(self, workbook):
		"""
		THIS METHOD IS YUCK, FIX ME
		When passed a workbook containing sheets of SBO object types
		and cells containing object names (column A) (and subtypes column B if subtypes)
		return a list of xml strings for each folder
		"""
		sheetnames = workbook.get_sheet_names()
		objects_xml = {}
		for sheetname in sheetnames:
			# loop through xl sheets to match sheetname against dict key
			elements = ''
			if sheetname in self.folder_object_types.keys():
				# if folder object types dict key value is list
				# need to consider seconf col to evaluate which list item
				sheet = workbook.get_sheet_by_name(sheetname)
				if type(self.folder_object_types[sheetname]) == list:
					for row in sheet.iter_rows():
						object_name = row[0].value
						object_type = self.get_object_subtype(sheetname, row[1].value)
						element = self.create_element_by_name_type(object_name, object_type)
						elements += '\n' + element
				else:
					# if folder object types dict key value is not list (ie a string)
					# just use the key bvalue as type
					for cell in sheet.columns[0]:
						if cell.value is not None:
							object_name = cell.value
							object_type = self.folder_object_types[sheetname]
							element = self.create_element_by_name_type(object_name, object_type)
							elements += '\n' + element
				objects_xml[sheetname] = elements
		
		return objects_xml
				
	
	def make_xml(self, write_result=True, print_result=False):	
		xml_child_folders = self.create_folders_from_list(self.child_folder_names)

		wb = openpyxl.load_workbook(xlfile)

		allsheetnames = wb.get_sheet_names()

		sheetnames = [s for s in allsheetnames if "meta" not in s]

		xml_str = self.xml_head

		application_folders = str()
		
		wb = openpyxl.load_workbook(xlfile)
		
		if self.commonobjects_xlfile is not None:
			wb_commonobjects = openpyxl.load_workbook(self.commonobjects_xlfile)
			for sheetname in sheetnames:
				# return dict of xml strings for each child
				# {'Variables': '<OI NAME=...", 'Alarms': '<OI NAME=...", ...}
				grandchildren = self.create_objects_from_excelbook(wb_commonobjects)
			xml_child_common_folders = self.create_folders_from_list(self.child_folder_names, grandchildren)
						
		# create common folder
		grandchildren = None
		application_folders += '\n' + self.create_folder_by_name('_Common', children=xml_child_common_folders)
		# loop thpugh xl sheets and create applications folders
		for sheetname in sheetnames:
			sheet = wb.get_sheet_by_name(sheetname)
			children = self.create_folders_from_excelsheet(sheet, children=xml_child_folders)
			application_folder = self.create_folder_by_name(sheetname, children=children)
			application_folders += '\n' + application_folder

		xml_str += '\n' + self.create_folder_by_name('Applications', application_folders)

		xml_str += '\n' +self.xml_foot
		
		if print_result:
			print(xml_str)
		if write_result:
			with open(xmlfile, "w") as outfile:
				outfile.write(xml_str)


# EXECUTE
if __name__ == "__main__":
	# declare filenames/paths here
	xlfile = 'Application Tree.xlsx'
	xmlfile = 'SBO Applications Tree.xml'
	objects_xlfile = 'ddc_objects.xlsx'
	
	# instantiate ApplicationTreeBuilder object and make xml
	apptreebuilder = ApplicationTreeBuilder(xlfile=xlfile, xmlfile=xmlfile, commonobjects_xlfile=objects_xlfile)
	apptreebuilder.make_xml(write_result=True)
