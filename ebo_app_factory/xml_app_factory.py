import openpyxl
from xml.dom import minidom
import sys

# <?xml version="1.0" encoding="UTF-8"?>
# <ObjectSet ExportMode="Special" Note="TypesFirst" Version="3.2.1.630">
#   <MetaInformation>
#     <ExportMode Value="Special"/>
#     <RuntimeVersion Value="3.2.1.630"/>
#     <SourceVersion Value="3.2.1.630"/>
#     <ServerFullPath Value="/ebo_app_factory"/>
#   </MetaInformation>
#   <Types>
#	 	...
#	</Types>
#	<ExportedObjects>
#		<OI NAME="VAV-L21-INT4" TYPE="system.base.Folder">
# 			...
#		</OI>
#	</ExportedObjects>
# </ObjectSet>


class ApplicationTemplate(object):

	def __init__(self, xml_in_file, print_result=False):
		'''
		creates a dictionary of lists of child DOM Elements from the template xml file
		where each key represents the tagname of each eleemnt that should live in the root element of the DOM

		ApplicationTemplate.template_child_elements_dict = {
			'Types': [<DOM Element: ObjectType at 0x7f5665bec700>, <DOM Element: ObjectType at 0x7f5665adef70>],
			'ExportedObjects': [<DOM Element: OI at 0x7f5665afe430>]
		}
		'''
		self.xml_in_doc = minidom.parse(xml_in_file)
		# list of nodes to export children
		self.template_nodes = [
			'Types',
			'ExportedObjects'
		]
		self.template_child_elements_dict = self.get_template_child_elements_dict()
		if print_result:
			for key, value in self.template_child_elements_dict.items():
				for element in value:
					print(key, value)
					print(element.toxml())

	def get_child_nodes_by_element_tagname(self, tagname, elements_only=False):
		'''
		assumes 1st element returned in getElementsByTagName
		returns list of child nodes inside Element 'tagname'
		'''
		if elements_only:
			return self.get_child_elements(self.xml_in_doc.getElementsByTagName(tagname)[0].childNodes)
		else:
			return self.xml_in_doc.getElementsByTagName(tagname)[0].childNodes

	def get_attr_if_exists(self, node, attr_name):
		if node.hasAttribute(attr_name):
			return node.getAttribute(attr_name)
		else:
			return None

	def get_child_elements(self, node):
		'''
		returns list of child nodes of type Node.ELEMENT_NODE only
		'''
		return [child for child in node if child.nodeType == minidom.Node.ELEMENT_NODE]

	def get_template_child_elements_dict(self):
		'''
		returns dictionary of lists of xml child element nodes required for xml app AppTemplate.
		Keys are found in self.template_nodes: Types, ExportedObjects
		'''
		template_child_elements_dict = {}
		for key in self.template_nodes:
				template_child_elements_dict[key] = self.get_child_nodes_by_element_tagname(key, elements_only=True)
		return template_child_elements_dict


class FactoryInputsFromSpreadsheet(object):

	def __init__(self, xlfile=None, sheetname=None, print_result=False):
		"""
		read in a spreadsheet containing a tables of:
		- app template placeholder substrings
		- app copy replacement substrings
		By default, all sheets are read except 'meta'. To limit which sheets are read, set sheetname to a string mathcing the sheet name or a list of strings equal to each sheet name required.
		each non empty cell of the first row of each sheet read is stored as a template placeholder substring key:value store
		each non empty second and subsequent row represents an app to be copied from the template, stored as a list of key:value stores,
		the cells of which correspond to replacement strings

		Example:

		Sheet1
		x		A		B		C
		1		VAV-1	Zn1		Room 2.31
		2		VAV-2	Zn2		Meeting Room 7
		3		VAV-3.2	Zn3B	Level 3 reception

		factory_placeholders = {
		'Sheet1A': 'VAV-1',
		'Sheet1B': 'Zn1',
		'Sheet1C': 'Room 2.31'
		}

		factory_placeholders_sorted = {
			'Sheet1': {
				'Sheet1A': 'VAV-1',
				'Sheet1B': 'Zn1',
				'Sheet1C': 'Room 2.31'
			}
		}

		factory_copy_substrings = [
			{
				'Sheet1A': 'VAV-2',
				'Sheet1B': 'Zn2',
				'Sheet1C': 'Meeting Room 7'
			},
			{
				'Sheet1A': 'VAV-3.2',
				'Sheet1B': 'Zn3B',
				'Sheet1C': 'Level 3 reception'
			},
		]

		factory_copy_substrings_sorted = {
			'Sheet1': [
				{
					'Sheet1A': 'VAV-2',
					'Sheet1B': 'Zn2',
					'Sheet1C': 'Meeting Room 7'
				},
				{
					'Sheet1A': 'VAV-3.2',
					'Sheet1B': 'Zn3B',
					'Sheet1C': 'Level 3 reception'
				},
			]
		}

		"""
		self.xlfile = xlfile
		self.show_progress = True
		self.create_factory_inputs_from_excel(sheetname=sheetname)
		if print_result:
			print(self.factory_placeholders)
			print(self.factory_copy_substrings)

	def create_factory_inputs_from_excel(self, sheetname=None):
		'''
		if sheetname not specified, all sheets will be read into one big list (except 'meta')
		if sheetname is str, only read the sheet with name sheetname
		if sheetname is list, read each member of list as a sheetname

		self.factory_copy_substrings is a flattened list of copy substrings
		factory_copy_substrings_sorted is a dict of lists of copy substrings,
		where each key represents a sheet
		'''

		workbook = openpyxl.load_workbook(self.xlfile, data_only=True)

		if sheetname == None:
			allsheetnames = workbook.sheetnames
			sheetnames = [s for s in allsheetnames if "meta" not in s]
		elif isinstance(sheetname, str):
			sheetnames = [sheetname]
		elif isinstance(sheetname, list):
			sheetnames = sheetname

		self.factory_placeholders = {}
		self.factory_placeholders_sorted = {}
		self.factory_copy_substrings = []
		self.factory_copy_substrings_sorted = {}

		if self.show_progress: print('\nCreating factory inputs from:', sheetnames)

		for sheetname in sheetnames:
			(placeholders, factory_copy_substrings) = self.create_factory_inputs_from_xl_sheet(sheetname, workbook)
			self.factory_placeholders.update(placeholders)
			self.factory_copy_substrings.extend(factory_copy_substrings)
			self.factory_placeholders_sorted[sheetname] = placeholders
			self.factory_copy_substrings_sorted[sheetname] = factory_copy_substrings

	def create_factory_inputs_from_xl_sheet(self, sheetname, workbook):

		sheet = workbook[sheetname]
		placeholders = {}
		factory_copy_substrings = []
		for row in sheet.iter_rows():
			factory_copy = {}
			first_row = True
			for cell in row:
				key = sheetname + cell.column_letter
				if cell.row == 1:
					placeholders[key] = cell.value
				else:
					first_row = False
					factory_copy[key] = cell.value
			if not first_row:
				factory_copy_substrings.append(factory_copy)
		return (placeholders, factory_copy_substrings)


class ApplicationFactory(object):

	def __init__(
		self,
		template_child_elements_dict=None,
		factory_placeholders=None,
		factory_copy_substrings=None,
		xml_out_file=None,
		ebo_version="3.2.1.630",
		ebo_server_full_path="/EBOApplicationFactory_v0.1",
		ebo_export_mode="Special",
		show_progress=True
	):
		self.show_progress = show_progress
		self.xml_out_file = xml_out_file
		self.template_child_elements_dict = template_child_elements_dict
		self.factory_placeholders = factory_placeholders
		self.factory_copy_substrings = factory_copy_substrings

		doc_template_str = """<?xml version="1.0" encoding="utf-8"?>
			<ObjectSet ExportMode="Special" Note="TypesFirst" Version="3.2.1.630">
			\t<MetaInformation>
			\t\t<ExportMode Value="Special" />
			\t\t<RuntimeVersion Value="3.2.1.630"/>
			\t\t <SourceVersion Value="3.2.1.630"/>
			\t\t<ServerFullPath Value="/ebo_app_factory"/>
			\t</MetaInformation>
			</ObjectSet>"""
		self.factory_doc = minidom.parseString(doc_template_str)
		self.doc_root_element_tagname = 'ObjectSet'

	def stdout_progress(self, step, total_steps):
		if self.show_progress:
			sys.stdout.write('\r')
			sys.stdout.write("%d%%" % (step/total_steps*100))
			sys.stdout.flush()
		return step+1

	def make_document(self, write_result=True, print_result=False):
		'''
		The xml document is constructed as follows:
		<ObjectSet>
			{{ header stuff }}
			<Types>
				{{ self.factory_copies_dict['Types'] }}
			</Types>
			<ExportedObjects>
				{{ self.factory_copies_dict['ExportedObjects'] }}
			</ExportedObjects>
		</ObjectSet>
		'''
		# check if factory copies has already been created
		if not hasattr(self, 'factory_copies_dict'):
			self.make_copies()
		# report progress
		if self.show_progress: print('\nCreating document...')
		size = 0
		for node in self.factory_copies_dict:
			size += len(self.factory_copies_dict[node])
		progress = 1
		# loop through dictionary keys for each element to insert children
		for node, elements in self.factory_copies_dict.items():
			# create an empty child element inside root DOM Element ObjectSet
			factory_element = self.factory_doc.createElement(node)
			self.factory_doc.documentElement.appendChild(factory_element)
			# self.doc_template.getElementsByTagName(self.doc_root_element_tagname)[0].appendChild()
			# loop through elements and insert as children
			for child_element in elements:
				self.factory_doc.getElementsByTagName(node)[0].appendChild(child_element)
				# report progress
				progress = self.stdout_progress(progress, size)
		if print_result:
			print(self.factory_doc.toprettyxml(encoding='utf-8'))
			print(self.factory_doc)
		if write_result:
			if self.show_progress: print('\nWriting document to "' + self.xml_out_file + '" ...')
			with open(self.xml_out_file, "wb") as outfile:
				outfile.write(self.factory_doc.toxml(encoding='utf-8'))
			if self.show_progress: print('\nDone.\n')


	def make_copies(self):
		'''
		self.template_child_elements_dict AND factory_copies_dict = {
			'Types': [
				<DOM Element: ObjectType at 0x7fac26249f70>,
				<DOM Element: ObjectType at 0x7fac26144ee0>
			],
			'ExportedObjects': [
				<DOM Element: OI at 0x7fac261653a0>
			]
		}
		self.factory_copy_substrings = [
			{'Sheet1A': 'VAV-L16-INT1', 'Sheet1B': 'L16-INT1'},
			{'Sheet1A': 'VAV-L16-INT10', 'Sheet1B': 'L16-INT10'},
			{'Sheet1A': 'VAV-L16-INT11', 'Sheet1B': 'L16-INT11'}
		]
		'''
		# report progress
		if self.show_progress: print('Creating copies...')
		size = len(self.factory_copy_substrings)
		progress = 1
		# create empty factory copies dictionary
		factory_copies_dict = {key: [] for key in self.template_child_elements_dict}
		# loop through copy strings list
		for copy_substrings in self.factory_copy_substrings:
			# loop through template child DOM Elements, find and replace placeholders
			# append copy DOM Element to factory copies dictionary of lists
			for node, elements in self.template_child_elements_dict.items():
				copy_elements = []
				for element in elements:
					copy_element = self.replace_placeholders(element, copy_substrings)
					factory_copies_dict[node].append(copy_element)
			# report progress
			progress = self.stdout_progress(progress, size)
		self.factory_copies_dict = factory_copies_dict

	def replace_placeholders(self, element, copy_substrings):
		'''
		find and replace xml element template placeholder strings with copy strings
		'''
		# convert DOM Element to xml string
		factory_copy_element_str = element.toxml()
		# find and replace placeholders with copy values
		for key, placeholder_value in self.factory_placeholders.items():
			factory_copy_element_str = factory_copy_element_str.replace(placeholder_value, copy_substrings[key])
		# convert xml string back to DOM Element
		factory_copy_element = minidom.parseString(factory_copy_element_str).getElementsByTagName(element.tagName)[0]
		# self.xml_in_doc.getElementsByTagName(tagname)[0]
		return factory_copy_element



class ApplicationFactoryManager(object):

	def __init__(
		self,
		template_map=None,
		xlfile=None,
		sheetname=None,
		xml_out_file_prefix=None,
		ebo_version="3.2.1.630",
		ebo_server_full_path="/EBOApplicationFactory_v0.1",
		ebo_export_mode="Special",
		show_progress=True
	):
		self.show_progress = show_progress
		self.xlfile = xlfile
		self.xml_out_file_prefix = xml_out_file_prefix
		self.template_map = template_map

		self.get_factory_inputs(sheetname=sheetname)
		self.get_app_templates()

	def get_app_templates(self):
		if self.show_progress: print('\nCreating template documents...')
		for group, items in self.template_map.items():
			 items['elements'] = ApplicationTemplate(items['templateFilename'], print_result=False).template_child_elements_dict


	def get_factory_inputs(self, sheetname=None):
		if self.show_progress: print('\nCreating factory inputs from workbook "'+self.xlfile+'"')
		self.factory_inputs = FactoryInputsFromSpreadsheet(self.xlfile, sheetname=sheetname, print_result=False)
		self.factory_placeholders_sorted = self.factory_inputs.factory_placeholders_sorted
		self.factory_copy_substrings_sorted = self.factory_inputs.factory_copy_substrings_sorted

	def make_documents(self):
		for group, factory_copy_substrings in self.factory_copy_substrings_sorted.items():
			# app_template.template_child_elements_dict
			if self.show_progress: print('\nStarting production on "'+group+'" applications...')
			app_factory = ApplicationFactory(
				template_child_elements_dict=self.template_map[group]['elements'],
				factory_placeholders=self.factory_placeholders_sorted[group],
				factory_copy_substrings=factory_copy_substrings,
				xml_out_file=self.xml_out_file_prefix+'_'+group+'.xml',
			)
			app_factory.make_document()

# EXECUTE
if __name__ == "__main__":

	########################
	# Basic example
	########################
	# declare filenames/paths here
	xl_in_file = 'examples/basic apps example.xlsx'
	xml_in_file = 'examples/VAV-L21-INT4 application special.xml'
	xml_out_file = 'examples/generated_ebo_apps_basic_example.xml'

	# instantiate AppTemplate object object
	app_template = ApplicationTemplate(xml_in_file, print_result=False)
	# instantiate FactoryInputsFromSpreadsheet object
	factory_inputs = FactoryInputsFromSpreadsheet(xl_in_file, print_result=False)
	# instantiate ApplicationFactory object and make xml
	app_factory = ApplicationFactory(
		template_child_elements_dict=app_template.template_child_elements_dict,
		factory_placeholders=factory_inputs.factory_placeholders,
		factory_copy_substrings=factory_inputs.factory_copy_substrings,
		xml_out_file=xml_out_file,
	)
	# app_factory.make_copies()
	app_factory.make_document()

	########################
	# Advanced example
	########################
	# declare filenames/paths here
	xl_sorted_in_file = 'examples/sorted apps example.xlsx'
	# create dictionary mapping Excel sheet names to template xml files
	template_map = {
	  'L2-3-All3StgHtg': {'templateFilename': 'examples/VAV-L21-NW2 application special.xml'},
	  'L4-12-3StgHtg': {'templateFilename': 'examples/VAV-L21-NW2 application special.xml'},
	  'L13-15-3StgHtg': {'templateFilename': 'examples/VAV-L21-NW2 application special.xml'},
	  'L16-32-3StgHtg': {'templateFilename': 'examples/VAV-L21-NW2 application special.xml'},
	  '1StgHtg': {'templateFilename': 'examples/VAV-L04-INT09 application special.xml'},
	  'L2-3NoHtg': {'templateFilename': 'examples/VAV-L21-INT4 application special.xml'},
	  'L4-12NoHtg': {'templateFilename': 'examples/VAV-L21-INT4 application special.xml'},
	  'L13-15NoHtg': {'templateFilename': 'examples/VAV-L21-INT4 application special.xml'},
	  'L16-32NoHtg': {'templateFilename': 'examples/VAV-L21-INT4 application special.xml'},
	}
	# instantiate ApplicationFactoryManager object
	app_factory_manager = ApplicationFactoryManager(
		template_map=template_map,
		xlfile=xl_sorted_in_file,
		xml_out_file_prefix='examples/example_ebo_apps',
	)
	# make xml files
	app_factory_manager.make_documents()
