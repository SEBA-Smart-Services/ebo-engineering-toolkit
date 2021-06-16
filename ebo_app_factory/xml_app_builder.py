import openpyxl
from xml.dom import minidom
import sys

# <?xml version="1.0" encoding="UTF-8"?>
# <ObjectSet ExportMode="Special" Note="TypesFirst" Version="3.2.1.630">
#   <MetaInformation>
#     <ExportMode Value="Special"/>
#     <RuntimeVersion Value="3.2.1.630"/>
#     <SourceVersion Value="3.2.1.630"/>
#     <ServerFullPath Value="/ebo_app_factory"/>
#   </MetaInformation>
#   <Types>
#	 	...
#	</Types>
#	<ExportedObjects>
#		<OI NAME="VAV-L21-INT4" TYPE="system.base.Folder">
# 			...
#		</OI>
#	</ExportedObjects>
# </ObjectSet>


class ApplicationTemplate(object):

	def __init__(self, xml_in_file, print_result=False):
		self.xml_in_doc = minidom.parse(xml_in_file)
		# list of nodes to export children
		self.template_nodes = [
			'Types',
			'ExportedObjects'
		]
		self.template_child_elements_dict = self.get_template_child_elements_dict()
		if print_result:
			for key, value in self.template_child_elements_dict.items():
				for element in value:
					print(key, value)
					print(element.toxml())
			print('\n')
			print(self.template_child_elements_dict)
			print('\n')

	def get_child_nodes_by_element_tagname(self, tagname, elements_only=False):
		'''
		assumes 1st element returned in getElementsByTagName
		returns list of child nodes inside Element 'tagname'
		'''
		if elements_only:
			return self.get_child_elements(self.xml_in_doc.getElementsByTagName(tagname)[0].childNodes)
		else:
			return self.xml_in_doc.getElementsByTagName(tagname)[0].childNodes

	def get_attr_if_exists(self, node, attr_name):
		if node.hasAttribute(attr_name):
			return node.getAttribute(attr_name)
		else:
			return None

	def get_child_elements(self, node):
		'''
		returns list of child nodes of type Node.ELEMENT_NODE only
		'''
		return [child for child in node if child.nodeType == minidom.Node.ELEMENT_NODE]

	def get_template_child_elements_dict(self):
		'''
		returns dictionary of lists of xml child element nodes required for xml app AppTemplate.
		Keys are found in self.template_nodes: Types, ExportedObjects
		'''
		template_child_elements_dict = {}
		for key in self.template_nodes:
				template_child_elements_dict[key] = self.get_child_nodes_by_element_tagname(key, elements_only=True)
		return template_child_elements_dict


class FactoryInputsFromSpreadsheet(object):

	def __init__(self, xlfile=None, print_result=False):
		"""
		read in a spreadsheet containing a tables of:
		- app template placeholder substrings
		- app copy replacement substrings
		all sheets are read except 'meta'
		each non empty cell of the first row of each sheet read is stored as a template placeholder substring key:value store
		each non empty second and subsequent row represents an app to be copied from the template, stored as a list of key:value stores,
		the cells of which correspond to replacement strings

		Example:

		Sheet1
		x		A		B		C
		1		VAV-1	Zn1		Room 2.31
		2		VAV-2	Zn2		Meeting Room 7
		3		VAV-3.2	Zn3B	Level 3 reception

		factory_placeholders = {
		'Sheet1A': 'VAV-1',
		'Sheet1B': 'Zn1',
		'Sheet1C': 'Room 2.31'
		}

		factory_copy_substrings = [
		{
		'Sheet1A': 'VAV-2',
		'Sheet1B': 'Zn2',
		'Sheet1C': 'Meeting Room 7'
		},
		{
		'Sheet1A': 'VAV-3.2',
		'Sheet1B': 'Zn3B',
		'Sheet1C': 'Level 3 reception'
		},
		]

		"""
		self.xlfile = xlfile
		self.create_factory_inputs_from_excel()
		if print_result:
			print(self.factory_placeholders)
			print(self.factory_copy_substrings)

	def create_factory_inputs_from_excel(self):

		workbook = openpyxl.load_workbook(self.xlfile)
		allsheetnames = workbook.sheetnames
		sheetnames = [s for s in allsheetnames if "meta" not in s]

		self.factory_placeholders = {}
		self.factory_copy_substrings = []

		for sheetname in sheetnames:
			# if folder object types dict key value is list
			# need to consider seconf col to evaluate which list item
			sheet = workbook[sheetname]
			for row in sheet.iter_rows():
				factory_copy = {}
				first_row = True
				for cell in row:
					key = sheetname + cell.column_letter
					if cell.row == 1:
						self.factory_placeholders[key] = cell.value
					else:
						first_row = False
						factory_copy[key] = cell.value
				if not first_row:
					self.factory_copy_substrings.append(factory_copy)


class ApplicationFactory(object):

	def __init__(
		self,
		template_child_elements_dict=None,
		factory_placeholders=None,
		factory_copy_substrings=None,
		xml_out_file=None,
		ebo_version="3.2.1.630",
		ebo_server_full_path="/EBOApplicationFactory_v0.1",
		ebo_export_mode="Special",
		show_progress=True
	):
		self.show_progress = show_progress
		self.xml_out_file = xml_out_file
		self.template_child_elements_dict = template_child_elements_dict
		self.factory_placeholders = factory_placeholders
		self.factory_copy_substrings = factory_copy_substrings

		doc_template_str = """<?xml version="1.0" encoding="utf-8"?>
			<ObjectSet ExportMode="Special" Note="TypesFirst" Version="3.2.1.630">
			\t<MetaInformation>
			\t\t<ExportMode Value="Special" />
			\t\t<RuntimeVersion Value="3.2.1.630"/>
			\t\t <SourceVersion Value="3.2.1.630"/>
			\t\t<ServerFullPath Value="/ebo_app_factory"/>
			\t</MetaInformation>
			</ObjectSet>"""
		self.factory_doc = minidom.parseString(doc_template_str)
		self.doc_root_element_tagname = 'ObjectSet'

		print(self.factory_doc.toprettyxml(encoding='utf-8'))
		print(self.factory_doc.getElementsByTagName(self.doc_root_element_tagname)[0])

	def stdout_progress(self, step, total_steps):
		if self.show_progress:
			sys.stdout.write('\r')
			sys.stdout.write("%d%%" % (step/total_steps*100))
			sys.stdout.flush()
		return step+1

	def make_document(self, write_result=True, print_result=False):
		'''
		The xml document is constructed as follows:
		<ObjectSet>
			{{ header stuff }}
			<Types>
				{{ self.factory_copies_dict['Types'] }}
			</Types>
			<ExportedObjects>
				{{ self.factory_copies_dict['ExportedObjects'] }}
			</ExportedObjects>
		</ObjectSet>
		'''
		# check if factory copies has already been created
		if not hasattr(self, 'factory_copies_dict'):
			self.make_copies()
		# report progress
		print('\nCreating document...')
		size = 0
		for node in self.factory_copies_dict:
			size += len(self.factory_copies_dict[node])
		progress = 1
		# loop through dictionary keys for each element to insert children
		for node, elements in self.factory_copies_dict.items():
			# create an empty child element inside root DOM Element ObjectSet
			factory_element = self.factory_doc.createElement(node)
			self.factory_doc.documentElement.appendChild(factory_element)
			# self.doc_template.getElementsByTagName(self.doc_root_element_tagname)[0].appendChild()
			# loop through elements and insert as children
			for child_element in elements:
				self.factory_doc.getElementsByTagName(node)[0].appendChild(child_element)
				# report progress
				progress = self.stdout_progress(progress, size)
		if print_result:
			print(self.factory_doc.toprettyxml(encoding='utf-8'))
			print(self.factory_doc)
		if write_result:
			print('\nWriting document to "' + self.xml_out_file + '" ...')
			with open(self.xml_out_file, "wb") as outfile:
				outfile.write(self.factory_doc.toxml(encoding='utf-8'))
			print('\nDone.\n')


	def make_copies(self):
		'''
		self.template_child_elements_dict AND factory_copies_dict = {
			'Types': [
				<DOM Element: ObjectType at 0x7fac26249f70>,
				<DOM Element: ObjectType at 0x7fac26144ee0>
			],
			'ExportedObjects': [
				<DOM Element: OI at 0x7fac261653a0>
			]
		}
		self.factory_copy_substrings = [
			{'Sheet1A': 'VAV-L16-INT1', 'Sheet1B': 'L16-INT1'},
			{'Sheet1A': 'VAV-L16-INT10', 'Sheet1B': 'L16-INT10'},
			{'Sheet1A': 'VAV-L16-INT11', 'Sheet1B': 'L16-INT11'}
		]
		'''
		# report progress
		print('\nCreating copies...')
		size = len(self.factory_copy_substrings)
		progress = 1
		# create empty factory copies dictionary
		factory_copies_dict = {key: [] for key in self.template_child_elements_dict}
		# loop through copy strings list
		for copy_substrings in self.factory_copy_substrings:
			# loop through template child DOM Elements, find and replace placeholders
			# append copy DOM Element to factory copies dictionary of lists
			for node, elements in self.template_child_elements_dict.items():
				copy_elements = []
				for element in elements:
					copy_element = self.replace_placeholders(element, copy_substrings)
					factory_copies_dict[node].append(copy_element)
			# report progress
			progress = self.stdout_progress(progress, size)
		self.factory_copies_dict = factory_copies_dict

	def replace_placeholders(self, element, copy_substrings):
		'''
		find and replace xml element template placeholder strings with copy strings
		'''
		# convert DOM Element to xml string
		factory_copy_element_str = element.toxml()
		# find and replace placeholders with copy values
		for key, placeholder_value in self.factory_placeholders.items():
			factory_copy_element_str = factory_copy_element_str.replace(placeholder_value, copy_substrings[key])
		# convert xml string back to DOM Element
		factory_copy_element = minidom.parseString(factory_copy_element_str).getElementsByTagName(element.tagName)[0]
		# self.xml_in_doc.getElementsByTagName(tagname)[0]
		return factory_copy_element


# EXECUTE
if __name__ == "__main__":
	# declare filenames/paths here
	xl_in_file = 'examples/apps.xlsx'
	xml_in_file = 'examples/VAV-L21-INT4 application special.xml'
	xml_out_file = 'examples/generated_ebo_apps.xml'
	objects_xlfile = None

	# instantiate AppTemplate object object
	app_template = ApplicationTemplate(xml_in_file, print_result=False)
	# instantiate FactoryInputsFromSpreadsheet object
	factory_inputs = FactoryInputsFromSpreadsheet(xl_in_file, print_result=False)
	# instantiate ApplicationFactory object and make xml
	app_factory = ApplicationFactory(
		template_child_elements_dict=app_template.template_child_elements_dict,
		factory_placeholders=factory_inputs.factory_placeholders,
		factory_copy_substrings=factory_inputs.factory_copy_substrings,
		xml_out_file=xml_out_file,
	)
	# app_factory.make_copies()
	app_factory.make_document()
